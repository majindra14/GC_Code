import pandas as pd
import numpy as np
from numpy import array
import matplotlib.pyplot as plt
import math
import openpyxl
from statistics import * 
from cycler import cycler
from scipy import stats
from datetime import datetime
import os
from openpyxl import load_workbook

class bcolors:
    OK = '\033[92m' #GREEN
    WARNING = '\033[93m' #YELLOW
    FAIL = '\033[91m' #RED
    RESET = '\033[0m' #RESET COLOR

    
def red(text):
    """
    Returns a string that is red when printed
    """
    return bcolors.FAIL + text + bcolors.RESET 


def yellow(text):
    """
    Returns a string that is yellow when printed
    """
    return bcolors.WARNING + text + bcolors.RESET


def green(text):
    """
    Returns a string that is green when printed
    """
    return bcolors.OK + text + bcolors.RESET
    

def load_data(file_list, sheet, remove_sample_list, remove_product_list, output = False):
    """
    This function gets all the corrected concentration data from each file in file_list, removes samples and products specified
    by remove_sample_list and remove_product_list, and arranges the data into a dataframe. A master_sample_list is also generated
    which contains the list of all samples in the dataframe, and df_dict holds the dataframes from each individual file, in case
    analysis needs to be done on single files. This function also outputs information detected from each file for troubleshooting.
    """
    # Make sure file_list is a list, even if it's a single string
    file_list = [file_list] if isinstance(file_list, str) else file_list
    # Create empty dataframe to hold all data
    master_df = pd.DataFrame()
    # Dictionary to hold miscillaenous information for review if desired
    df_dict = {}
    # List to hold all sample names for review if desired. Also used to filter duplicate samples across multiple spreadsheets
    master_sample_list = []
    
    # Create progress bar
    if not output:
        print('Progress: |', end = '')
    
    for file in file_list:
        # Get the corrected concentration data
        df, sample_list = getCorrectedConcentration(file, sheet)
        # Remove any samples that are in remove_sample_list
        df_removed = df[~df['Sample Group'].isin(remove_sample_list)]
        # Remove any samples that have already been added from other sheets
        df_removed_dup = df_removed[~df_removed['Sample Group'].isin(master_sample_list)]
        # Identify which were removed due to remove_sample_list
        removed = list(set(df['Sample Group']) & set(remove_sample_list))
        # Identify which were removed due to duplicates
        duplicates = list(set(df_removed['Sample Group']) & set(master_sample_list))
        # Identify which were kept after removal from remove_sample_list (not duplicates)
        sample_list_kept = list(df_removed['Sample Group'].unique())
        
        # Update dataframe dictionary
        df_dict[file] = df
        # Update master dataframe
        master_df = master_df.append(df_removed).reset_index(drop = True)
        
        # Remove products that are in remove_product_list
        master_df = master_df[~master_df['Products'].isin(remove_product_list)]
        
        
        # Print if output is True
        if output:
            color_product_list = [red(product) if product in remove_product_list else product for product in df['Products'].unique()]
            color_sample_list = []
            for sample in list(df['Sample Group'].unique()):
                if sample in remove_sample_list:
                    color_sample_list.append(red(sample))
                    continue
                if sample in master_sample_list:
                    color_sample_list.append(yellow(sample))
                    continue
                color_sample_list.append(green(sample))
                
            print('\033[1m' + file + '\033[0m')
            print('Detected products:      [', end = '')
            for i in range(len(color_product_list) - 1):
                print('\'' + color_product_list[i] + '\'', end = ', ')
            print('\'' + color_product_list[-1] + '\'', end = ']\n')
            print('Detected sample groups: [', end = '') 
            for i in range(len(color_sample_list) - 1):
                print('\'' + color_sample_list[i] + '\'', end = ', ')
            print('\'' + color_sample_list[-1] + '\'', end = ']\n')
        else:
            print('#', end = '')
        # Update master sample list
        master_sample_list.extend(sample_list_kept)
    
    if not output:
        print('|')
        
    print('\033[1m' + 'Done!' + '\033[0m')
    return master_df, master_sample_list, df_dict


def getCorrectedConcentration(file, sheet):
    """
    Filters the data in the <sheet> worksheet in the <file> excel file. Use for 'Corrected Concentration' worksheet.
    Returns dataframe with filtered Corrected Concentration data and additional columns for grouping data by sample and product.
    """
    # Read excel file
    wb = openpyxl.load_workbook(file)
    df = pd.read_excel(file, sheet_name=sheet)
    
    # Get sample names from first column by filtering for strings only and not counting rows with 'Peak#'. Rename to 'Sample Names'
    sample_names = df[pd.to_numeric(df[0], errors='coerce').isnull()][0][df[0] != 'Peak#'].dropna().to_frame().rename(columns={0:'Sample Names'})
    # Add sample names column to the original dataframe and rename column 5 (which contains the product IDs) to 'Products'
    df = pd.concat([df,sample_names],axis=1).rename(columns={5:'Products'})
    # Replace '_' character with ':' for propper naming convention
    df['Products'] = df['Products'].apply(lambda product : product.replace('_',':'))
    # Forward fill the sample names
    df['Sample Names'] = df['Sample Names'].fillna(method = 'ffill')
    # Get rid of number identifier (expected format is <sample name>-<sample number>). Removes the last item separated by a hyphen
    df['Sample Group'] = df['Sample Names'].apply(lambda sample_name: "-".join(sample_name.split('-')[:-1]))
    # Create column to identify which file the data came from
    df['Source File'] = file
    # Get rid of blanks in 'Corrected Concentration'
    df = df.dropna(subset=['Corrected Concentration'])
    return df, list(df['Sample Group'].unique())


def dataParse(df, sample_order = None, product_order = None):
    """
    Uses cleaned Corrected Concentration dataframe to calculate triplicate averages (scaled and absolute) and standard errors
    Returns three dataframes (average, scaled, and standard error) with rows corresponding to samples and columns corresponding to products
    """
    # Calculate means, scaled means, and standard errors using pivot table-like operation
    df_mean = pd.pivot_table(df, values = 'Corrected Concentration', index = ['Sample Group'], columns = ['Products'])
    df_sum = list(df_mean.sum(axis = 1))
    df_scaled = df_mean.divide(df_sum, axis='index')
    df_err = pd.pivot_table(df, values = 'Corrected Concentration', index = ['Sample Group'], columns = ['Products'], aggfunc = stats.sem)
    
    # If sample order is provided, then change order
    if sample_order != None:
        sample_order = [sample for sample in sample_order if sample in list(df['Sample Group'])]
        df_mean = df_mean.reindex(index = sample_order)
        df_scaled = df_scaled.reindex(index = sample_order)
        df_err = df_err.reindex(index = sample_order)
    # If product order is provided, then change order
    if product_order != None:
        product_order = [product for product in product_order if product in list(df['Products'])]
        df_mean = df_mean[product_order]
        df_scaled = df_scaled[product_order]
        df_err = df_err[product_order]
        
    return df_mean, df_scaled, df_err


def summary_stats(df, sample_order = None, product_order = None):
    """
    Generates summary statistics of product distributions for each sample given dataframe of clean Corrected Concentrations
    Returns dataframe containing percentage of total titer of each product for each sample group, and a final column containing total titer averages
    """
    df = df.copy(True)
    # Calculate the total titer for each sample
    tot_titer = df.groupby('Sample Names').sum()
    # Label the sample group for each sample
    tot_titer['Sample Group'] = tot_titer.index.to_series().apply(lambda sample_name: "-".join(sample_name.split('-')[:-1]))
    # Create new column in original dataframe containing the total titer for each data point
    df['Total Titer'] = df['Sample Names'].apply(lambda sample_name: tot_titer.loc[sample_name]['Corrected Concentration'])
    # Create new column to calculate the percentage that each product contributes to the total titer
    df['Percent of Total Titer'] = df['Corrected Concentration']/df['Total Titer']*100
    
    # Calculate means and standard errors using pivot table-like operation
    df_pct = pd.pivot_table(df, values = 'Percent of Total Titer', index = ['Sample Group'], columns = ['Products'])
    df_pct_err = pd.pivot_table(df, values = 'Percent of Total Titer', index = ['Sample Group'], columns = ['Products'], aggfunc = stats.sem)
    
    # Change data-type to string to allow string combination of '±' sign
    sample_mean = df_pct.applymap(('{:,.' + str(3) + 'g}').format).astype(str)
    sample_err = df_pct_err.applymap(('{:,.' + str(3) + 'g}').format).astype(str)
    sample_stats = sample_mean + '±' + sample_err

    # Calculate mean and standard error of total titer for each sample group and change data-type to string to allow string combination of '±' sign
    total_mean = tot_titer.groupby('Sample Group').first()['Corrected Concentration'].map(('{:,.' + str(3) + 'g}').format).astype(str)
    total_err = tot_titer.groupby('Sample Group').sem()['Corrected Concentration'].map(('{:,.' + str(3) + 'g}').format).astype(str)
    total_stats = (total_mean + '±' + total_err).to_frame().rename(columns = {'Corrected Concentration':'mg/L'})
    
    # Get dataframe of just source files
    source_files = df.drop_duplicates(subset = 'Sample Group').set_index('Sample Group')['Source File']
    
    # Combine total titer data with sample data and return summary_stats dataframe
    summary_stats = pd.concat([sample_stats, total_stats, source_files], axis = 1)
    
    # If sample order is provided, then change order
    if sample_order != None:
        sample_order_filt = [sample for sample in sample_order if sample in sample_stats.index]
        summary_stats = summary_stats.reindex(index = sample_order_filt)
    # If product order is provided, then change order
    if product_order != None:
        product_order_filt = [product for product in product_order if product in sample_stats] + ['mg/L'] + ['Source File']
        summary_stats = summary_stats[product_order_filt]
    
    # Update column names with '%' symbol
    summary_stats.columns = ['% ' + str(product) if product != 'Source File' and product != 'mg/L' else str(product) for product in summary_stats.columns]
    
    return summary_stats, total_stats, source_files, init_stats


def set_hatch(ax, chains):
    """
    Takes an axis object and sets the hatch property so that the presence of hatches alternates
    """
    bars = ax.patches
    # List of hatch patterns for each bar
    patterns = (['']*len(ax.get_xticks()) + ['\\']*len(ax.get_xticks()))*len(chains)
    # Set the hatch patterns
    for bar, pattern in zip(bars, patterns):
        bar.set_hatch(pattern)
    return patterns


def save_df_to_excel(df, filename, sheet_name, replace = False):
    """
    Saves a dataframe to an Excel file with name <filename> in sheet <sheet_name>. Several warnings
    """
    # Get file path and check if it exists
    file_path = os.path.join(os.getcwd(), filename)
    file_exists = os.path.exists(file_path)
    
    # If the file doesnt exist, or if the user wants to replace the existing file, create the file, but check one more time with user.
    if not file_exists or replace:
        if replace and file_exists:
            proceed = ('Existing file will be replaced. Proceed? Yes/No: ')
            if proceed.lower() in ['yes','ye','y']:
                df.to_excel(filename, sheet_name = sheet_name)
            else:
                print('Operation cancelled')
                return
        else:
            df.to_excel(filename, sheet_name = sheet_name)            
        print('Done!')
    # Otherwise...
    else:
        # Generate workbook
        wb = load_workbook(file_path)
        # Generate writer engine
        writer = pd.ExcelWriter(file_path, engine = 'openpyxl')
        # Assigning the workbook to the writer engine
        writer.book = wb
        
        # If the sheet already exists in the file
        if sheet_name in wb.sheetnames:
            # Prompt user to see if they want to replace the existing sheet
            print('The sheet name \'' + sheet_name + '\' already exists in ' + filename + '.')
            proceed = input('Existing data in \'' + sheet_name + '\' will be replaced with new data. Proceed? Yes/No: ')
            # If they want to replace the existing sheet, then do it
            if proceed.lower() in ['yes','ye','y']:
                df.to_excel(writer, sheet_name = sheet_name)
                print('Done!')
            # Otherwise, do nothing
            else:
                print('Operation cancelled.')
        # If sheet doesn't exist, create it and save the dataframe to it.
        else:
            df.to_excel(writer, sheet_name = sheet_name)
            print('New sheet \'' + sheet_name + '\', added to ' + filename + '\nDone!')
        # Save and close writer
        writer.save()
        writer.close()